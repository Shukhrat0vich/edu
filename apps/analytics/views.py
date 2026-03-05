"""Analytics views for EduAnalytics."""
import json
import io
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.utils.decorators import method_decorator
from django.views import View
from django.http import HttpResponse
from django.contrib.admin.views.decorators import staff_member_required

from .services import AnalyticsService
from apps.users.models import UserRole


def role_required(*roles):
    """Decorator to restrict view access by role."""
    from functools import wraps
    from django.shortcuts import redirect

    def decorator(view_func):
        @wraps(view_func)
        def _wrapped_view(request, *args, **kwargs):
            if not request.user.is_authenticated:
                return redirect('users:login')
            if request.user.role not in roles:
                return HttpResponse('<h2>403 Forbidden — Access Denied</h2>', status=403)
            return view_func(request, *args, **kwargs)
        return _wrapped_view
    return decorator


@method_decorator(login_required, name='dispatch')
class AnalyticsOverviewView(View):
    """Full analytics overview — admin only."""
    template_name = 'analytics/overview.html'

    def get(self, request):
        if request.user.role not in [UserRole.ADMIN, UserRole.TEACHER]:
            return HttpResponse('403 Forbidden', status=403)

        subject_id = request.GET.get('subject')
        faculty = request.GET.get('faculty')

        df = AnalyticsService.get_grades_dataframe(subject_id=subject_id, faculty=faculty)
        stats = AnalyticsService.compute_descriptive_stats(df)
        subject_performance = AnalyticsService.get_subject_performance()
        top_students = AnalyticsService.get_top_students(limit=10)
        score_dist = AnalyticsService.get_score_distribution()

        context = {
            'stats': stats,
            'subject_performance': subject_performance,
            'top_students': top_students,
            'score_distribution_json': json.dumps(score_dist),
            'subject_perf_json': json.dumps(subject_performance),
        }
        return render(request, self.template_name, context)


@method_decorator(login_required, name='dispatch')
class CorrelationView(View):
    """Correlation matrix view."""
    template_name = 'analytics/correlation.html'

    def get(self, request):
        if request.user.role not in [UserRole.ADMIN, UserRole.TEACHER]:
            return HttpResponse('403 Forbidden', status=403)

        corr = AnalyticsService.get_correlation_matrix()
        context = {
            'correlation_json': json.dumps(corr),
            'correlation': corr,
        }
        return render(request, self.template_name, context)


@method_decorator(login_required, name='dispatch')
class HeatmapView(View):
    """Performance heatmap view — admin only."""
    template_name = 'analytics/heatmap.html'

    def get(self, request):
        if request.user.role != UserRole.ADMIN:
            return HttpResponse('403 Forbidden', status=403)

        heatmap_data = AnalyticsService.get_heatmap_data()
        context = {
            'heatmap_json': json.dumps(heatmap_data),
        }
        return render(request, self.template_name, context)


@method_decorator(login_required, name='dispatch')
class ExportExcelView(View):
    """Export grade data to Excel."""

    def get(self, request):
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return HttpResponse('openpyxl not installed', status=500)

        df = AnalyticsService.get_grades_dataframe()
        subject_perf = AnalyticsService.get_subject_performance()
        top_students = AnalyticsService.get_top_students(50)

        wb = openpyxl.Workbook()

        # Sheet 1: All Grades
        ws1 = wb.active
        ws1.title = 'All Grades'
        header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        if not df.empty:
            headers = list(df.columns)
            for col_idx, header in enumerate(headers, 1):
                cell = ws1.cell(row=1, column=col_idx, value=header.replace('_', ' ').title())
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            for row_idx, row in enumerate(df.itertuples(index=False), 2):
                for col_idx, value in enumerate(row, 1):
                    ws1.cell(row=row_idx, column=col_idx, value=value)

        # Sheet 2: Subject Performance
        ws2 = wb.create_sheet('Subject Performance')
        s2_headers = ['Subject', 'Avg Score', 'Max Score', 'Min Score', 'Std Dev', 'Pass Rate %', 'Students']
        for col_idx, h in enumerate(s2_headers, 1):
            cell = ws2.cell(row=1, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font

        for row_idx, subj in enumerate(subject_perf, 2):
            ws2.cell(row=row_idx, column=1, value=subj['subject_name'])
            ws2.cell(row=row_idx, column=2, value=subj['avg_score'])
            ws2.cell(row=row_idx, column=3, value=subj['max_score'])
            ws2.cell(row=row_idx, column=4, value=subj['min_score'])
            ws2.cell(row=row_idx, column=5, value=subj['std_score'])
            ws2.cell(row=row_idx, column=6, value=subj['pass_rate'])
            ws2.cell(row=row_idx, column=7, value=subj['total_students'])

        # Sheet 3: Top Students
        ws3 = wb.create_sheet('Top Students')
        s3_headers = ['Rank', 'Name', 'Faculty', 'Group', 'Avg Score', 'GPA']
        for col_idx, h in enumerate(s3_headers, 1):
            cell = ws3.cell(row=1, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font

        for row_idx, st in enumerate(top_students, 2):
            ws3.cell(row=row_idx, column=1, value=st['rank'])
            ws3.cell(row=row_idx, column=2, value=st['full_name'])
            ws3.cell(row=row_idx, column=3, value=st['faculty'])
            ws3.cell(row=row_idx, column=4, value=st['group'])
            ws3.cell(row=row_idx, column=5, value=st['avg_score'])
            ws3.cell(row=row_idx, column=6, value=st['gpa'])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="edu_analytics_report.xlsx"'
        return response


@method_decorator(login_required, name='dispatch')
class ExportPDFView(View):
    """Export analytics report to PDF."""

    def get(self, request):
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        except ImportError:
            return HttpResponse('reportlab not installed', status=500)

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []

        # Title
        elements.append(Paragraph('EduAnalytics — Performance Report', styles['Title']))
        elements.append(Spacer(1, 12))

        # Summary stats
        summary = AnalyticsService.get_admin_summary()
        summary_data = [
            ['Metric', 'Value'],
            ['Total Students', summary['total_students']],
            ['Total Teachers', summary['total_teachers']],
            ['Average GPA', summary['avg_gpa']],
            ['At-Risk Students', summary['at_risk_count']],
            ['Avg Attendance %', summary['avg_attendance']],
        ]
        t = Table(summary_data, colWidths=[200, 200])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F4F6')]),
        ]))
        elements.append(Paragraph('Summary Statistics', styles['Heading2']))
        elements.append(Spacer(1, 8))
        elements.append(t)
        elements.append(Spacer(1, 20))

        # Top students
        elements.append(Paragraph('Top 10 Students', styles['Heading2']))
        elements.append(Spacer(1, 8))
        top_students = AnalyticsService.get_top_students(10)
        if top_students:
            ts_data = [['Rank', 'Name', 'Faculty', 'Avg Score', 'GPA']]
            for s in top_students:
                ts_data.append([s['rank'], s['full_name'], s['faculty'], s['avg_score'], s['gpa']])
            t2 = Table(ts_data, colWidths=[40, 160, 120, 80, 60])
            t2.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10B981')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#ECFDF5')]),
            ]))
            elements.append(t2)

        doc.build(elements)
        buffer.seek(0)

        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="edu_analytics_report.pdf"'
        return response
